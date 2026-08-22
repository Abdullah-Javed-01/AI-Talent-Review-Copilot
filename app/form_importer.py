import argparse
import re
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd


FIELD_ALIASES = {
    "submitted_at": ["timestamp", "submitted at", "submission time"],
    "candidate_id": ["candidate id", "candidate_id", "application id", "applicant id"],
    "name": ["name", "full name", "candidate name", "applicant name"],
    "email": ["email", "email address", "e-mail"],
    "phone": ["phone", "phone number", "contact number", "mobile number", "contact"],
    "university": ["university", "university name", "institute", "institution"],
    "semester": [
        "semester",
        "current semester",
        "current semester status",
        "education status",
        "academic status",
    ],
    "applied_role": ["applied role", "role", "internship role", "position", "position applied for"],
    "current_city": ["current city", "city", "current location", "location"],
    "lahore_address": ["lahore address", "address in lahore", "address", "current address"],
    "shift_available": [
        "shift available",
        "working hours",
        "5pm-2am availability",
        "5 pm - 2 am availability",
        "5pm to 2am",
        "required shift",
        "are you available for 5 00 pm 2 00 am working hours",
    ],
    "onsite_available": [
        "onsite available",
        "onsite availability",
        "available onsite",
        "on-site availability",
        "work onsite",
        "available to work onsite in lahore",
        "are you available to work onsite in lahore",
    ],
    "linkedin_url": ["linkedin", "linkedin url", "linkedin profile", "linkedin profile url"],
    "resume_source": [
        "resume",
        "cv",
        "resume/cv",
        "resume cv",
        "upload resume",
        "upload cv",
        "resume link",
        "resume cv file upload",
        "resume cv file upload file",
    ],
}


def normalize_header(value):
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_yes_no(value):
    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "available", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "not available", "0"}:
        return "No"
    return str(value).strip()


def is_url(value):
    try:
        parsed = urlparse(str(value).strip())
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except ValueError:
        return False


def find_column(columns, aliases):
    normalized = {normalize_header(column): column for column in columns}

    # Prefer exact normalized matches first.
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized:
            return normalized[key]

    # Google Forms often turns labels into full questions. Match meaningful
    # alias phrases inside those questions without relying on exact punctuation.
    for alias in aliases:
        key = normalize_header(alias)
        if len(key) < 6:
            continue
        for normalized_header, original_header in normalized.items():
            if key in normalized_header:
                return original_header

    return None


def _read_export(input_path):
    input_path = Path(input_path)
    suffix = input_path.suffix.lower()

    if suffix == ".csv":
        # Keep values as strings so phone numbers such as 0309... retain their
        # leading zero and no form value is silently converted to a number.
        return pd.read_csv(input_path, dtype=str, keep_default_na=False), {}

    if suffix in {".xlsx", ".xlsm"}:
        raw = pd.read_excel(input_path, dtype=str, keep_default_na=False)
        hyperlinks = {}

        # XLSX exports preserve Google Sheets hyperlinks whereas CSV exports
        # keep only the displayed filename. Capture those links as provenance.
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(input_path, data_only=False, read_only=False)
            sheet = workbook[workbook.sheetnames[0]]
            headers = [cell.value for cell in sheet[1]]

            for row_offset, excel_row in enumerate(sheet.iter_rows(min_row=2), start=0):
                for column_index, cell in enumerate(excel_row):
                    if cell.hyperlink and column_index < len(headers):
                        header = headers[column_index]
                        hyperlinks[(row_offset, header)] = cell.hyperlink.target
        except Exception:
            hyperlinks = {}

        return raw, hyperlinks

    raise ValueError("Supported form exports are CSV and XLSX files.")


def import_form_export(input_file, output_csv):
    input_file = Path(input_file)
    output_csv = Path(output_csv)

    raw, hyperlinks = _read_export(input_file)
    normalized = pd.DataFrame(index=raw.index)

    resolved_columns = {}
    for internal_name, aliases in FIELD_ALIASES.items():
        source_column = find_column(raw.columns, aliases)
        resolved_columns[internal_name] = source_column
        normalized[internal_name] = (
            raw[source_column].fillna("").astype(str).str.strip()
            if source_column
            else ""
        )

    if normalized["candidate_id"].eq("").all():
        normalized["candidate_id"] = [
            f"FORM-{index:06d}"
            for index in range(1, len(normalized) + 1)
        ]
    else:
        missing_ids = normalized["candidate_id"].eq("")
        for offset, row_index in enumerate(normalized.index[missing_ids], start=1):
            normalized.loc[row_index, "candidate_id"] = f"FORM-MISSING-{offset:06d}"

    normalized["shift_available"] = normalized["shift_available"].map(normalize_yes_no)
    normalized["onsite_available"] = normalized["onsite_available"].map(normalize_yes_no)

    resume_column = resolved_columns.get("resume_source")
    normalized["resume_display_name"] = normalized["resume_source"]

    # Prefer the actual hyperlink preserved in an XLSX export. CSV exports
    # cannot retain the underlying Drive link and therefore keep the displayed
    # filename as the source reference.
    if resume_column:
        for row_index in normalized.index:
            hyperlink = hyperlinks.get((row_index, resume_column))
            if hyperlink:
                normalized.loc[row_index, "resume_source"] = hyperlink

    normalized["resume_filename"] = normalized["resume_display_name"].map(
        lambda value: "" if not value or is_url(value) else Path(str(value)).name
    )
    normalized["resume_source_type"] = normalized["resume_source"].map(
        lambda value: "remote_url" if is_url(value) else "local_reference" if value else "missing"
    )

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    normalized.to_csv(output_csv, index=False)

    return {
        "rows": len(normalized),
        "output_csv": str(output_csv),
        "resolved_columns": resolved_columns,
        "unresolved_fields": [
            key for key, value in resolved_columns.items()
            if value is None and key not in {"candidate_id", "submitted_at"}
        ],
        "drive_links_preserved": int(normalized["resume_source"].map(is_url).sum()),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Normalize an exported internship application form into the Talent Review Copilot schema."
    )
    parser.add_argument("input_file", help="Path to the exported form CSV or XLSX")
    parser.add_argument(
        "--output",
        default="data/applications/imported_applications.csv",
        help="Normalized CSV output path",
    )
    args = parser.parse_args()

    result = import_form_export(args.input_file, args.output)

    print(f"Imported {result['rows']} application(s).")
    print(f"Normalized CSV: {result['output_csv']}")

    if result["unresolved_fields"]:
        print("Unresolved optional/form fields:")
        for field in result["unresolved_fields"]:
            print(f"- {field}")

    if result["drive_links_preserved"]:
        print(
            f"\nPreserved {result['drive_links_preserved']} remote resume link(s) from the export. "
            "Authorized Google Drive access is still required to retrieve those files automatically."
        )


if __name__ == "__main__":
    main()
